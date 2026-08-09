# -*- coding: utf-8 -*-
"""최근 N영업일 분석 결과 → 그래프 3종(PNG) + 엑셀 원본. (기간은 collect.py의 NDAYS)"""
import os, sys, json
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(HERE))
import pdf_change as P
from analyze import flows_aum, net_buys, DAYS, EOK

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm

# 한글 폰트
_av = {f.name for f in fm.fontManager.ttflist}
for _f in ("Malgun Gothic", "NanumGothic", "AppleGothic", "Noto Sans CJK KR"):
    if _f in _av:
        matplotlib.rcParams["font.family"] = _f; break
matplotlib.rcParams["axes.unicode_minus"] = False

# 검증된 팔레트 (light)
SURF="#fcfcfb"; INK="#0b0b0b"; INK2="#52514e"; MUTED="#898781"; GRID="#e1e0d9"; BASE="#c3c2b7"
# 다이버징: 매수=빨강, 매도=파랑 (국내 관행) — 팔레트의 blue↔red 쌍
BUY="#e34948"; SELL="#2a78d6"
CAT=["#2a78d6","#eb6834","#1baf7a","#eda100","#e87ba4","#008300","#4a3aa7","#e34948"]

def md(ymd): return f"{ymd[4:6]}/{ymd[6:]}"

def style(ax, title=None, ylab=None):
    ax.set_facecolor(SURF)
    for s in ("top","right"): ax.spines[s].set_visible(False)
    for s in ("left","bottom"): ax.spines[s].set_color(BASE); ax.spines[s].set_linewidth(0.8)
    ax.grid(True, axis="y", color=GRID, lw=0.7, zorder=0)
    ax.set_axisbelow(True)
    ax.tick_params(colors=MUTED, labelsize=8.5, length=0)
    if title: ax.set_title(title, fontsize=10.5, fontweight="bold", color=INK, loc="left", pad=6)
    if ylab: ax.set_ylabel(ylab, fontsize=8.5, color=INK2)

def zeroline(ax):
    ax.axhline(0, color=BASE, lw=0.9, zorder=1)

def end_labels(ax, items, fontsize=8):
    """라인 끝 라벨을 겹치지 않게 배치. items=[(y값, 텍스트)]"""
    if not items: return
    lo, hi = ax.get_ylim()
    gap = (hi - lo) * 0.052
    it = sorted(items, key=lambda t: t[0])
    ys = [v for v, _ in it]
    for i in range(1, len(ys)):                      # 아래→위로 밀기
        if ys[i] - ys[i-1] < gap: ys[i] = ys[i-1] + gap
    over = ys[-1] - (hi - gap*0.3)
    if over > 0:                                     # 위로 넘치면 전체 내림
        ys = [y - over for y in ys]
        for i in range(len(ys)-2, -1, -1):
            if ys[i+1] - ys[i] < gap: ys[i] = ys[i+1] - gap
    x = ax.get_xlim()[1]
    for (orig, txt), y in zip(it, ys):
        ax.annotate(txt, (x, y), xytext=(-4, 0), textcoords="offset points",
                    fontsize=fontsize, color=INK2, va="center", ha="right")

def bar_labels_h(ax, vals, fmt="{:+,.0f}억", fontsize=8):
    """가로막대 값 라벨 — 축 밖으로 넘치면 막대 안쪽(흰색)에 배치."""
    lo, hi = ax.get_xlim()
    span = hi - lo
    for i, v in enumerate(vals):
        outside = (v >= 0 and v + span*0.13 < hi) or (v < 0 and v - span*0.13 > lo)
        if outside:
            ax.annotate(fmt.format(v), (v, i), textcoords="offset points",
                        xytext=(5 if v >= 0 else -5, 0), ha="left" if v >= 0 else "right",
                        va="center", fontsize=fontsize, fontweight="bold", color=INK)
        else:
            ax.annotate(fmt.format(v), (v, i), textcoords="offset points",
                        xytext=(-6 if v >= 0 else 6, 0), ha="right" if v >= 0 else "left",
                        va="center", fontsize=fontsize, fontweight="bold", color="white")

# ── ① 자금유출입 + AUM (패시브 | 액티브 소형다중, 이중축 금지) ──
def chart1(agg, out):
    days = sorted({r["일자"] for r in agg})
    D = {("패시브",),("액티브",)}
    def series(typ, key):
        m = {r["일자"]: r[key] for r in agg if r["구분"] == typ}
        return [m.get(d, 0.0)/EOK for d in days]
    fig, axes = plt.subplots(3, 2, figsize=(13.5, 10.5), dpi=170, facecolor=SURF)
    fig.suptitle(f"① 코스닥 ETF 자금 유출입(설정·환매)과 순자산  ·  {days[0][:4]}-{md(days[0])} ~ {md(days[-1])} ({len(days)}영업일)",
                 fontsize=13, fontweight="bold", color=INK, x=0.012, ha="left", y=0.985)
    x = range(len(days))
    for col, typ in enumerate(["패시브", "액티브"]):
        fl = series(typ, "순유출입(원)"); au = series(typ, "순자산(원)")
        cum = []; s = 0.0
        for v in fl: s += v; cum.append(s)
        # 1행: 일간 순유출입
        ax = axes[0][col]; style(ax, f"{typ} — 일간 순유출입", "억원")
        ax.bar(x, fl, color=[BUY if v >= 0 else SELL for v in fl], width=0.68, zorder=2)
        zeroline(ax)
        # 2행: 누적 순유출입
        ax = axes[1][col]; style(ax, f"{typ} — 누적 순유출입", "억원")
        ax.plot(x, cum, color=BUY if cum[-1] >= 0 else SELL, lw=2, zorder=2)
        ax.fill_between(x, 0, cum, color=(BUY if cum[-1] >= 0 else SELL), alpha=0.12, zorder=1)
        zeroline(ax)
        ax.annotate(f"{cum[-1]:+,.0f}억", (len(days)-1, cum[-1]), textcoords="offset points",
                    xytext=(-4, 6), ha="right", fontsize=9.5, fontweight="bold", color=INK)
        # 3행: 순자산(AUM)
        ax = axes[2][col]; style(ax, f"{typ} — 순자산 총액", "억원")
        ax.plot(x, au, color=CAT[0], lw=2, zorder=2)
        ax.annotate(f"{au[-1]:,.0f}억", (len(days)-1, au[-1]), textcoords="offset points",
                    xytext=(-4, 6), ha="right", fontsize=9.5, fontweight="bold", color=INK)
        for row in range(3):
            a = axes[row][col]
            a.set_xticks(list(x)[::2]); a.set_xticklabels([md(days[i]) for i in list(x)[::2]], rotation=45, ha="right")
    fig.text(0.012, 0.005, "순유출입 = 상장좌수 변화 × NAV (주가등락 제외한 순수 설정/환매)  ·  빨강=유입, 파랑=유출  ·  좌우 패널은 축 스케일이 다름",
             fontsize=8, color=MUTED)
    fig.tight_layout(rect=[0, 0.015, 1, 0.965])
    fig.savefig(out, facecolor=SURF); plt.close(fig)
    return out

def coverage_note(rows, skipped):
    """차트 하단에 넣을 커버리지 문구 (결측 ETF 명시)."""
    used = sorted({r["ETF"] for r in rows})
    from collections import Counter
    miss = Counter(s[0] for s in skipped if s[2] == "PDF없음")
    base = f"대상: 액티브 {len(used)}종"
    if miss:
        m = ", ".join(f"{k.split()[0]} {k.split()[-1][:6]}({v}일)" for k, v in miss.items())
        return base + f"  ·  ⚠ 데이터 결측 제외: {m}"
    return base + " (전종 완비)"

def _agg_by(rows, key):
    """{key: {ymd: 금액(억)}}"""
    d = defaultdict(lambda: defaultdict(float))
    for r in rows: d[r[key]][r["일자"]] += r["순매수금액(원)"]/EOK
    return d

def _topn(d, n):
    tot = {k: sum(v.values()) for k, v in d.items()}
    return sorted(tot, key=lambda k: -abs(tot[k]))[:n], tot

def chart2(rows, out, note="", order=None, ncol=3, cmp_days=None):
    """테마 보유비중 변화 — 덤벨(수준+변화 동시) + 소형다중(비중 추이, 각자 축)."""
    from analyze import theme_weights
    days, W, H = theme_weights()
    i0 = 0 if not cmp_days else max(0, len(days) - 1 - cmp_days)   # 덤벨 비교 시작점
    win = set(days[i0 + 1:])                                        # %p와 동일 구간의 순매수만 집계
    flow = defaultdict(float)
    for r in rows:
        if r["일자"] in win: flow[r["테마"]] += r["순매수금액(원)"] / EOK
    groups = [g for g in (order or sorted(W[-1], key=lambda k: -W[-1].get(k, 0))) if g in W[-1]]
    n = len(groups); nrow = (n + ncol - 1) // ncol
    fig = plt.figure(figsize=(13.5, 3.8 + nrow * 2.25), dpi=170, facecolor=SURF)
    gs = fig.add_gridspec(1 + nrow, ncol, height_ratios=[1.6] + [1.0] * nrow, hspace=0.62, wspace=0.26)
    nchg = len(days) - 1 - i0
    fig.suptitle(f"② 액티브 ETF 테마별 보유비중 변화  ·  {days[i0][:4]}-{md(days[i0])} → {md(days[-1])} ({nchg}영업일 변화)",
                 fontsize=13, fontweight="bold", color=INK, x=0.012, ha="left", y=0.985)
    # (1) 덤벨 — 시작 비중 → 끝 비중 (수준과 변화를 한 번에)
    ax = fig.add_subplot(gs[0, :]); style(ax, f"보유비중: {md(days[i0])} → {md(days[-1])}", None)
    items = sorted(groups, key=lambda g: W[-1].get(g, 0))
    for i, g in enumerate(items):
        a, b = W[i0].get(g, 0), W[-1].get(g, 0)
        col = BUY if b >= a else SELL
        ax.plot([a, b], [i, i], color=col, lw=2.5, zorder=2, solid_capstyle="round")
        ax.scatter([a], [i], s=42, color=SURF, edgecolor=MUTED, lw=1.4, zorder=3)   # 시작(빈 점)
        ax.scatter([b], [i], s=64, color=col, zorder=4)                              # 현재(채운 점)
        ax.annotate(f"{b:.1f}%  ({b-a:+.1f}%p)", (max(a, b), i), textcoords="offset points",
                    xytext=(9, 0), va="center", fontsize=9, fontweight="bold", color=INK)
    ax.set_yticks(range(len(items))); ax.set_yticklabels(items, fontsize=10, color=INK)
    ax.grid(True, axis="x", color=GRID, lw=0.7); ax.grid(False, axis="y")
    hi = max(max(W[i0].get(g, 0), W[-1].get(g, 0)) for g in groups)
    ax.set_xlim(-1, hi * 1.30)
    ax.set_xlabel("보유비중(%)      ○ 시작   ● 현재      빨강=비중 확대, 파랑=축소", fontsize=8.5, color=INK2)
    # (2) 소형다중 — 테마별 비중 추이(각자 축)
    x = list(range(len(days)))
    for i, g in enumerate(groups):
        ax = fig.add_subplot(gs[1 + i // ncol, i % ncol])
        c = [W[j].get(g, 0) for j in x]
        col = BUY if c[-1] >= c[i0] else SELL
        style(ax, f"{g}  {c[-1]:.1f}%  ({c[-1]-c[i0]:+.1f}%p)", None)
        ax.plot(x, c, color=col, lw=2, zorder=2)
        ax.fill_between(x, min(c) - (max(c)-min(c))*0.25 if max(c) > min(c) else 0, c,
                        color=col, alpha=0.12, zorder=1)
        ax.set_xticks(x[::3]); ax.set_xticklabels([md(days[j]) for j in x[::3]], rotation=45, ha="right", fontsize=8)
        ax.tick_params(labelsize=8)
        ax.annotate(f"순매수 {flow.get(g, 0):+,.0f}억", (0.03, 0.06), xycoords="axes fraction",
                    fontsize=8, color=MUTED)   # %p와 같은 구간
    fig.text(0.012, 0.005,
             "보유비중 = 테마 보유액 ÷ 액티브 4종 주식 보유액 합계  ·  매매(리밸런싱)와 주가등락이 함께 반영된 실제 포지션 변화\n"
             "(설정/환매는 전 종목이 비례 증감해 비중에 영향 없음)  ·  아래 패널은 각자 축\n" + note,
             fontsize=8, color=MUTED)
    fig.savefig(out, facecolor=SURF, bbox_inches="tight"); plt.close(fig)
    return out


def chart3(rows, out, topn=15, note=""):
    days = sorted({r["일자"] for r in rows})
    bystk = _agg_by(rows, "종목")
    tot = {k: sum(v.values()) for k, v in bystk.items()}
    buys = sorted(tot.items(), key=lambda kv: -kv[1])[:topn]
    sells = sorted(tot.items(), key=lambda kv: kv[1])[:topn]
    fig = plt.figure(figsize=(13.5, 11.5), dpi=170, facecolor=SURF)
    gs = fig.add_gridspec(2, 2, height_ratios=[1.5, 1.0], hspace=0.3, wspace=0.32)
    fig.suptitle(f"③ 액티브 ETF 종목별 순매수금액  ·  {days[0][:4]}-{md(days[0])} ~ {md(days[-1])} ({len(days)}영업일)",
                 fontsize=13, fontweight="bold", color=INK, x=0.012, ha="left", y=0.985)
    for col, (data, lab, color) in enumerate([(buys, f"순매수 상위 {topn}", BUY), (sells, f"순매도 상위 {topn}", SELL)]):
        ax = fig.add_subplot(gs[0, col]); style(ax, f"{len(days)}영업일 누적 {lab}", None)
        names = [k for k, v in data][::-1]; vals = [v for k, v in data][::-1]
        ax.barh(range(len(names)), vals, color=color, height=0.72, zorder=2)
        ax.set_yticks(range(len(names))); ax.set_yticklabels(names, fontsize=9, color=INK)
        ax.axvline(0, color=BASE, lw=0.9)
        ax.grid(True, axis="x", color=GRID, lw=0.7); ax.grid(False, axis="y")
        lo, hi = min(vals + [0]), max(vals + [0]); pad = (hi - lo) * 0.16
        ax.set_xlim(lo - pad, hi + pad)
        bar_labels_h(ax, vals)
        ax.set_xlabel("억원", fontsize=8.5, color=INK2)
    # 하단: 상위 4매수 + 4매도 누적 추이
    ax = fig.add_subplot(gs[1, :]); style(ax, "주요 종목 누적 순매수 추이 (매수 상위4 · 매도 상위4)", "억원")
    x = list(range(len(days)))
    pick = [k for k, v in buys[:4]] + [k for k, v in sells[:4]]
    labs = []
    for i, nm in enumerate(pick):
        c = []; acc = 0.0
        for d in days: acc += bystk[nm].get(d, 0.0); c.append(acc)
        ax.plot(x, c, color=CAT[i % len(CAT)], lw=2, zorder=2, label=nm)
        labs.append((c[-1], f"{nm} {c[-1]:+,.0f}"))
    zeroline(ax)
    ax.set_xticks(x[::2]); ax.set_xticklabels([md(days[i]) for i in x[::2]], rotation=45, ha="right")
    ax.set_xlim(-0.5, len(days)+6.5)
    end_labels(ax, labs)
    fig.text(0.012, 0.005,
             "순매수금액 = 1CU 수량변화 × 종가 × 전체CU수  ·  빨강=순매수, 파랑=순매도\n" + note,
             fontsize=8, color=MUTED)
    fig.savefig(out, facecolor=SURF, bbox_inches="tight"); plt.close(fig)
    return out

def excel(agg, detail, rows, out):
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    def sheet(name, cols, data):
        ws = wb.create_sheet(name)
        ws.append(cols)
        for c in ws[1]: c.font = Font(bold=True)
        for r in data: ws.append([r.get(k) for k in cols])
        ws.freeze_panes = "A2"
        for i, k in enumerate(cols, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(11, min(24, len(str(k))+6))
        return ws
    wb.remove(wb.active)
    # 1) 그룹 자금흐름
    sheet("①자금흐름_집계", ["일자","구분","순유출입(원)","순자산(원)"], agg)
    sheet("①자금흐름_ETF별", ["일자","구분","ETF","코드","좌수","좌수변화","NAV","순유출입(원)","순자산(원)","순자산출처"], detail)
    # 2) 섹터 일별
    bysec = defaultdict(lambda: defaultdict(float))
    for r in rows: bysec[r["섹터"]][r["일자"]] += r["순매수금액(원)"]
    days = sorted({r["일자"] for r in rows})
    # 2-0) 테마 보유비중 일별
    from analyze import theme_weights
    wdays, W, H = theme_weights()
    ws = wb.create_sheet("②테마_보유비중(%)")
    ws.append(["테마"] + wdays + ["변화(%p)"])
    for c in ws[1]: c.font = Font(bold=True)
    for t in sorted(W[-1], key=lambda k: -W[-1].get(k, 0)):
        vals = [round(W[i].get(t, 0), 2) for i in range(len(wdays))]
        ws.append([t] + vals + [round(vals[-1] - vals[0], 2)])
    ws.freeze_panes = "B2"; ws.column_dimensions["A"].width = 14
    ws2 = wb.create_sheet("②테마_보유액(원)")
    ws2.append(["테마"] + wdays)
    for c in ws2[1]: c.font = Font(bold=True)
    for t in sorted(H[-1], key=lambda k: -H[-1].get(k, 0)):
        ws2.append([t] + [round(H[i].get(t, 0)) for i in range(len(wdays))])
    ws2.freeze_panes = "B2"; ws2.column_dimensions["A"].width = 14
    # 2-A) 테마 일별
    byth = defaultdict(lambda: defaultdict(float))
    for r in rows: byth[r["테마"]][r["일자"]] += r["순매수금액(원)"]
    days0 = sorted({r["일자"] for r in rows})
    ws = wb.create_sheet("②테마_일별순매수(원)")
    ws.append(["테마"] + days0 + ["합계"])
    for c in ws[1]: c.font = Font(bold=True)
    for s in sorted(byth, key=lambda k: -sum(byth[k].values())):
        vals = [round(byth[s].get(d, 0.0)) for d in days0]
        ws.append([s] + vals + [sum(vals)])
    ws.freeze_panes = "B2"; ws.column_dimensions["A"].width = 14
    # 2-B) 업종 일별
    ws = wb.create_sheet("②B업종_일별순매수(원)")
    ws.append(["섹터"] + days + ["합계"])
    for c in ws[1]: c.font = Font(bold=True)
    for s in sorted(bysec, key=lambda k: -sum(bysec[k].values())):
        vals = [round(bysec[s].get(d, 0.0)) for d in days]
        ws.append([s] + vals + [sum(vals)])
    ws.freeze_panes = "B2"; ws.column_dimensions["A"].width = 22
    # 3) 종목 일별
    bystk = defaultdict(lambda: defaultdict(float))
    stksec = {}; stkth = {}
    for r in rows:
        bystk[r["종목"]][r["일자"]] += r["순매수금액(원)"]
        stksec[r["종목"]] = r["섹터"]; stkth[r["종목"]] = r["테마"]
    ws = wb.create_sheet("③종목_일별순매수(원)")
    ws.append(["종목","테마","업종"] + days + ["합계"])
    for c in ws[1]: c.font = Font(bold=True)
    for s in sorted(bystk, key=lambda k: -sum(bystk[k].values())):
        vals = [round(bystk[s].get(d, 0.0)) for d in days]
        ws.append([s, stkth.get(s, ""), stksec.get(s, "")] + vals + [sum(vals)])
    ws.freeze_panes = "D2"
    ws.column_dimensions["A"].width = 18; ws.column_dimensions["B"].width = 12; ws.column_dimensions["C"].width = 20
    # 4) 원본 행
    sheet("원본_ETF종목일별", ["일자","ETF","종목","테마","섹터","수량변화(1CU)","종가","전체CU","순매수금액(원)"], rows)
    wb.save(out)
    return out

if __name__ == "__main__":
    agg, detail = flows_aum()
    rows, skipped = net_buys()
    ND = len({r["일자"] for r in agg})          # 실제 변동일수 → 파일명에 반영
    o1 = chart1(agg, os.path.join(HERE, f"①_자금유출입_AUM_{ND}일.png"))
    note = coverage_note(rows, skipped)
    from themes import ORDER as TORDER
    o2 = chart2(rows, os.path.join(HERE, f"②_테마별_비중변화_{ND}일.png"), note, order=TORDER)
    o3 = chart3(rows, os.path.join(HERE, f"③_종목별_순매수_{ND}일.png"), note=note)
    o4 = excel(agg, detail, rows, os.path.join(HERE, f"{ND}영업일_분석.xlsx"))
    for o in (o1, o2, o3, o4): print("생성:", os.path.basename(o))
    if skipped:
        from collections import Counter
        print("결측:", Counter(s[0] for s in skipped).most_common())
