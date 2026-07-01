# -*- coding: utf-8 -*-
"""
운용사 직접 당일 PDF → 전일대비 구성종목 수량 변화(PDF 변화) 계산. (액티브 ETF 전용)
두 가지 fetcher 모드:
  - dateapi : 날짜 파라미터 API. 임의 영업일 조회 가능 → 첫 실행부터 diff (예: KoAct/삼성액티브)
  - snapshot: 당일치만 주는 소스. 매 영업일 스냅샷 저장 → 2영업일째부터 self-diff (예: TIME)
스냅샷은 snapshots.json 에 {etf_key:{YYYYMMDD:{종목:수량}}} 누적.
curl 차단 사이트(PLUS/UNICORN/RISE/DB)는 헤드드 Selenium fetcher로 추가 예정(PC 켜둠).
"""
import sys, os, csv, json, requests
from datetime import datetime, date, timedelta

BASE = os.path.dirname(os.path.abspath(__file__))
SNAP_FILE = os.path.join(BASE, "snapshots.json")
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0 Safari/537.36"

HOLIDAYS_2026 = {
    "2026-01-01","2026-02-16","2026-02-17","2026-02-18","2026-03-02","2026-05-05",
    "2026-05-25","2026-06-03","2026-08-17","2026-09-24","2026-09-25","2026-10-05",
    "2026-10-09","2026-12-25","2026-12-31",
}
def is_trading_day(d): return d.weekday() < 5 and d.strftime("%Y-%m-%d") not in HOLIDAYS_2026
def prev_trading_day(d):
    p = d - timedelta(days=1)
    while not is_trading_day(p): p -= timedelta(days=1)
    return p

SKIP = ("현금","원화예금","원화 예금","예수금","설정현금","CASH","원화","달러","USD","선물","원화현금")
MIN_CHANGE = 3  # 이 미만(±1·±2주)의 수량확대/축소는 노이즈로 숨김. 신규편입·편출은 크기 무관 항상 표시.

# ── 스냅샷 저장소 ──────────────────────────────────────────────
def load_snap():
    if os.path.exists(SNAP_FILE):
        try: return json.load(open(SNAP_FILE, encoding="utf-8"))
        except Exception: pass
    return {}
def save_snap(s):
    tmp = SNAP_FILE + ".tmp"
    json.dump(s, open(tmp, "w", encoding="utf-8"), ensure_ascii=False)
    os.replace(tmp, SNAP_FILE)

# ── fetcher: 삼성액티브(KoAct) 날짜 API → {종목:수량}, 기준일 ──
def fetch_samsungactive(prod_id, ymd):
    g = f"{ymd[:4]}.{ymd[4:6]}.{ymd[6:]}"
    r = requests.get(f"https://www.samsungactive.co.kr/api/v1/product/etf-pdf/{prod_id}.do",
                     params={"gijunYMD": g},
                     headers={"User-Agent": UA, "Referer": f"https://www.samsungactive.co.kr/etf/view.do?id={prod_id}"},
                     timeout=20)
    d = r.json().get("pdf", {})
    out = {}
    for it in d.get("list", []):
        nm = (it.get("secNm") or "").strip()
        if not nm or any(s in nm for s in SKIP): continue
        try:
            q = float(str(it.get("applyQ", "0")).replace(",", ""))
            ev = float(str(it.get("evalA", "0")).replace(",", ""))
        except ValueError: continue
        if nm in out: out[nm][0] += q
        else: out[nm] = [q, (ev / q if q else None)]
    return out, str(d.get("gijunYMD", "")).replace(".", "")

# ── fetcher: 타임폴리오(TIME) pdf_excel.php (pdfDate 날짜 파라미터) → {종목:수량}, 기준일 ──
def fetch_time(idx, ymd):
    import openpyxl, io
    pdfdate = f"{ymd[:4]}-{ymd[4:6]}-{ymd[6:]}"  # YYYY-MM-DD (대시 포함이어야 함)
    r = requests.get("https://timeetf.co.kr/pdf_excel.php",
                     params={"idx": idx, "pdfDate": pdfdate},
                     headers={"User-Agent": UA, "Referer": f"https://timeetf.co.kr/m11_view.php?idx={idx}"},
                     timeout=20)
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    out = {}
    if rows:
        hdr = [str(c).strip() if c else "" for c in rows[0]]
        try:
            i_nm, i_q = hdr.index("종목명"), hdr.index("수량")
        except ValueError:
            return {}, ""
        i_ev = next((j for j, c in enumerate(hdr) if "평가금액" in c), None)
        for row in rows[1:]:
            if not row or len(row) <= max(i_nm, i_q): continue
            nm = str(row[i_nm]).strip() if row[i_nm] else ""
            if not nm or any(s in nm for s in SKIP): continue
            try:
                q = float(str(row[i_q]).replace(",", ""))
                ev = float(str(row[i_ev]).replace(",", "")) if i_ev is not None and row[i_ev] is not None else 0.0
            except (ValueError, TypeError): continue
            if nm in out: out[nm][0] += q
            else: out[nm] = [q, (ev / q if (q and ev) else None)]
    return out, (ymd if out else "")  # 파일에 기준일 셀 없음 → 요청한 날짜로

# ── fetcher: 한화(PLUS) 엑셀 다운로드 (날짜 파라미터) → {종목:수량}, 기준일 ──
def fetch_plus(n, ymd):
    import openpyxl, io
    r = requests.get("https://www.plusetf.co.kr/excel/product/pdf",
                     params={"n": n, "d": ymd, "title": "PLUS"},
                     headers={"User-Agent": UA, "Referer": f"https://www.plusetf.co.kr/product/detail?n={n}"},
                     timeout=20)
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    actual = ""
    hdr_i = None
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        line = " ".join(cells)
        if not actual:
            import re as _re
            m = _re.search(r"20\d\d[.\-]\d\d[.\-]\d\d", line)
            if m: actual = m.group().replace("-", "").replace(".", "")
        if "종목명" in cells and any("수량" in c for c in cells):
            hdr_i = i
            i_nm = cells.index("종목명")
            i_q = next(j for j, c in enumerate(cells) if "수량" in c)
            break
    out = {}
    if hdr_i is not None:
        for row in rows[hdr_i + 1:]:
            if not row or len(row) <= max(i_nm, i_q): continue
            nm = str(row[i_nm]).strip() if row[i_nm] else ""
            if not nm or any(s in nm for s in SKIP): continue
            try: q = float(str(row[i_q]).replace(",", ""))
            except (ValueError, TypeError): continue
            if nm in out: out[nm][0] += q
            else: out[nm] = [q, None]   # PLUS 엑셀엔 평가금액 없음 → 종가 미상
    return out, actual

# ── fetcher: KB(RISE) 구성종목 엑셀(HTML-xls, searchDate) → {종목:수량}, 기준일 ──
def fetch_rise(stid, ymd):
    import pandas as pd, io
    d = f"{ymd[:4]}-{ymd[4:6]}-{ymd[6:]}"
    r = requests.get("https://www.riseetf.co.kr/prod/finder/productViewTabExcel3",
                     params={"searchTargetId": stid, "searchDate": d},
                     headers={"User-Agent": UA, "Accept-Language": "ko-KR",
                              "Referer": f"https://www.riseetf.co.kr/prod/finderDetail/{stid}?searchFlag=viewtab2"},
                     timeout=20)
    out = {}
    try:
        t = pd.read_html(io.StringIO(r.text))[0]
    except Exception:
        return {}, ""
    hdr = None; i_ev = None
    for i in range(len(t)):
        vals = [str(x).strip() for x in t.iloc[i].tolist()]
        if "종목명" in vals and any("수량" in v for v in vals):
            i_nm = vals.index("종목명"); i_q = next(j for j, v in enumerate(vals) if "수량" in v)
            i_ev = next((j for j, v in enumerate(vals) if "평가" in v), None); hdr = i; break
    if hdr is None:
        return {}, ""
    for i in range(hdr + 1, len(t)):
        vals = [str(x).strip() for x in t.iloc[i].tolist()]
        if len(vals) <= max(i_nm, i_q): continue
        nm = vals[i_nm]
        if not nm or nm == "nan" or any(s in nm for s in SKIP): continue
        try:
            q = float(vals[i_q].replace(",", ""))
            ev = float(vals[i_ev].replace(",", "")) if i_ev is not None else 0.0
        except ValueError: continue
        if nm in out: out[nm][0] += q
        else: out[nm] = [q, (ev / q if (q and ev) else None)]
    return out, (ymd if out else "")

# ── fetcher: 미래에셋(TIGER) pdfListAjax.ajax (세션+fixDate 점형식) → {종목:수량}, 기준일 ──
def fetch_tiger(isin, ymd):
    import pandas as pd, io
    d = f"{ymd[:4]}.{ymd[4:6]}.{ymd[6:]}"  # 점 형식 필수
    base = "https://investments.miraeasset.com/tigeretf/ko/product/search/detail"
    s = requests.Session(); s.headers.update({"User-Agent": UA, "Accept-Language": "ko-KR"})
    s.get(f"{base}/index.do", params={"ksdFund": isin}, timeout=20)  # 세션 쿠키
    body = {"ksdFund": isin, "pageIndex": 1, "firstIndex": 0, "listCnt": 300,
            "fixDate": d, "prfPrd": "Week01", "order": "SRD"}
    r = s.post(f"{base}/pdfListAjax.ajax", data=body,
               headers={"X-Requested-With": "XMLHttpRequest", "Referer": f"{base}/index.do?ksdFund={isin}"}, timeout=20)
    out = {}
    try:
        df = pd.read_html(io.StringIO("<table>" + r.text + "</table>"))[0]
    except Exception:
        return {}, ""
    for _, row in df.iterrows():
        nm = str(row[1]).strip()  # col0=종목코드,col1=종목명,col2=수량,col3=평가금액,col4=비중
        if not nm or nm == "nan" or any(sk in nm for sk in SKIP): continue
        try:
            q = float(str(row[2]).replace(",", ""))
            ev = float(str(row[3]).replace(",", ""))
        except (ValueError, TypeError): continue
        if nm in out: out[nm][0] += q
        else: out[nm] = [q, (ev / q if (q and ev) else None)]
    return out, (ymd if out else "")

# ── fetcher: 현대(UNICORN) /api/etfPdf (fundCode,etfCode,ymd) → {종목:수량}, 기준일 ──
def fetch_unicorn(combo, ymd):
    fund, etf = combo.split(":")
    r = requests.get("https://www.hyundaiam.com/api/etfPdf",
                     params={"fundCode": fund, "etfCode": etf, "ymd": ymd},
                     headers={"User-Agent": UA, "Accept-Language": "ko-KR",
                              "Referer": "https://www.hyundaiam.com/kor/HD-KP-FG/HD-KP-FG-07-D.html"},
                     timeout=20)
    out = {}; actual = ""
    for it in r.json():
        actual = str(it.get("date", "")) or actual
        nm = (it.get("구성종목명") or "").strip()
        if not nm or any(s in nm for s in SKIP): continue
        try: out[nm] = out.get(nm, 0.0) + float(it.get("구성종목수", 0) or 0)
        except (ValueError, TypeError): pass
    return out, (actual or (ymd if out else ""))

# ── fetcher: DB(마이티) — 운용사 사이트는 TOP5만 → WiseReport(전체 수량). snapshot ──
def fetch_db(code):
    import sys as _sys
    SIB = r"C:/Users/정호준/Desktop/코스닥 etf 분류"
    if SIB not in _sys.path: _sys.path.insert(0, SIB)
    from etf_discovery import get_etf_pdf_from_wisereport
    df, d = get_etf_pdf_from_wisereport(code, return_date=True)
    out = {}
    if df is not None and not df.empty:
        for _, row in df.iterrows():
            nm = str(row.get("종목명", "")).strip()
            if not nm or nm == "nan" or any(s in nm for s in SKIP): continue
            try: out[nm] = out.get(nm, 0.0) + float(row.get("수량(주)", 0) or 0)
            except (ValueError, TypeError): pass
    return out, (d or "").replace("-", "")

# ── 등록된 액티브 ETF ──────────────────────────────────────────
# krx: KRX 단축코드(상장좌수 조회용), cu: 설정단위(CU) 크기 → 전체CU수 = 좌수/cu
ETFS = [
    {"name":"KoAct 코스닥액티브","am":"삼성액티브","mode":"dateapi","fetch":fetch_samsungactive,"id":"2ETFU6","krx":"0163Y0","cu":50000},
    {"name":"KoAct 바이오헬스케어액티브","am":"삼성액티브","mode":"dateapi","fetch":fetch_samsungactive,"id":"2ETFJ9","krx":"462900","cu":50000},
    {"name":"TIME 코스닥액티브","am":"타임폴리오","mode":"dateapi","fetch":fetch_time,"id":"24","krx":"0162Y0","cu":50000},
    {"name":"TIME K바이오액티브","am":"타임폴리오","mode":"dateapi","fetch":fetch_time,"id":"13","krx":"463050","cu":50000},
    {"name":"PLUS 코스닥150액티브","am":"한화","mode":"dateapi","fetch":fetch_plus,"id":"006399","krx":"0166N0","cu":50000},
    {"name":"RISE 바이오TOP10액티브","am":"KB","mode":"dateapi","fetch":fetch_rise,"id":"44I0","krx":"0000Z0","cu":50000},
    {"name":"TIGER 기술이전바이오액티브","am":"미래에셋","mode":"dateapi","fetch":fetch_tiger,"id":"KR70168K0008","krx":"0168K0","cu":10000},  # CU=10000 (좌수/CU 정수검증으로 확정; 기존 config 20000은 오류)
    # 제외(2026-06-01, 사용자 요청): UNICORN 포스트IPO(현대, 111188:476000) / 마이티(DB, 0001P0, WiseReport).
    #   fetch_unicorn·fetch_db 함수는 코드에 유지 → 재추가 가능.
]

# ── 상장좌수(KRX) → 전체 CU수 계산 ─────────────────────────────
_SHARES_CACHE = {}
_NAV_CACHE = {}   # {KRX단축코드: 순자산총액(원)} — 비중 분모(전체 ETF 액수)
_KRX_PX_CACHE = {}   # {종목명: 종가} — 운용사가 평가금액 미제공(예: PLUS)일 때 종가 fallback
def _krx_key():
    k = os.getenv("KRX_API_KEY", "")
    if k: return k
    # 로컬: sibling .env 에서 로드 (클라우드는 환경변수/시크릿 사용)
    envp = r"C:/Users/정호준/Desktop/코스닥 etf 분류/.env"
    try:
        for line in open(envp, encoding="utf-8", errors="replace"):
            if line.strip().startswith("KRX_API_KEY"):
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    except OSError:
        pass
    return ""

def listed_shares():
    """{KRX단축코드: 상장좌수} — 최근 영업일. 키 없거나 실패 시 빈 dict(전체금액 생략)."""
    if _SHARES_CACHE: return _SHARES_CACHE
    key = _krx_key()
    if not key: return {}
    for off in range(1, 8):
        d = (datetime.now().date() - timedelta(days=off)).strftime("%Y%m%d")
        try:
            r = requests.get("http://data-dbg.krx.co.kr/svc/apis/etp/etf_bydd_trd",
                             params={"AUTH_KEY": key, "basDd": d}, timeout=20)
            items = r.json().get("OutBlock_1", []) if r.status_code == 200 else []
        except Exception:
            items = []
        if items:
            for it in items:
                code = str(it.get("ISU_SRT_CD") or it.get("ISU_CD") or "")
                try: _SHARES_CACHE[code] = int(str(it.get("LIST_SHRS", "0")).replace(",", "") or 0)
                except ValueError: pass
                try: _NAV_CACHE[code] = int(str(it.get("INVSTASST_NETASST_TOTAMT", "0")).replace(",", "") or 0)
                except ValueError: pass
            break
    return _SHARES_CACHE

def num_cu(e):
    """ETF의 전체 CU수 = 상장좌수/CU. 좌수 미상이면 0."""
    sh = listed_shares().get(e.get("krx", ""), 0)
    cu = e.get("cu", 0)
    return (sh / cu) if (sh and cu) else 0

def etf_navtotal(e):
    """ETF 순자산총액(원, KRX 공식) = 비중 분모. 미상이면 0."""
    listed_shares()  # 캐시 채움(좌수와 동일 API)
    return _NAV_CACHE.get(e.get("krx", ""), 0)

def krx_kosdaq_prices():
    """{종목명: 종가} — KRX 코스닥 최근 영업일 종가. 운용사 평가금액 미제공 종목의 종가 fallback."""
    if _KRX_PX_CACHE: return _KRX_PX_CACHE
    key = _krx_key()
    if not key: return {}
    for off in range(1, 8):
        d = (datetime.now().date() - timedelta(days=off)).strftime("%Y%m%d")
        try:
            r = requests.get("http://data-dbg.krx.co.kr/svc/apis/sto/ksq_bydd_trd",
                             params={"AUTH_KEY": key, "basDd": d}, timeout=20)
            items = r.json().get("OutBlock_1", []) if r.status_code == 200 else []
        except Exception:
            items = []
        if items:
            for it in items:
                try: _KRX_PX_CACHE[it["ISU_NM"].strip()] = float(str(it["TDD_CLSPRC"]).replace(",", ""))
                except (ValueError, KeyError): pass
            break
    return _KRX_PX_CACHE

def _qp(v):
    """맵 값 → (수량, 종가). 값은 [수량,종가](신형) 또는 숫자(구형) 모두 허용."""
    if isinstance(v, (list, tuple)):
        return float(v[0]), (float(v[1]) if v[1] not in (None, "") else None)
    return float(v), None

def diff(tmap, pmap):
    rows = []
    for nm in set(tmap) | set(pmap):
        tq, tp = _qp(tmap.get(nm, (0.0, None)))
        pq, pp = _qp(pmap.get(nm, (0.0, None)))
        ch = tq - pq
        if pq == 0 and tq > 0:   typ = "신규편입"
        elif tq == 0 and pq > 0: typ = "편출"
        elif ch > 0:             typ = "수량확대"
        elif ch < 0:             typ = "수량축소"
        else:                    typ = "유지"
        price = tp if tp else pp          # 당일 종가 우선, 편출이면 전일 종가
        if not price:                     # 운용사 평가금액 미제공(PLUS 등) → KRX 코스닥 종가로 보완
            price = krx_kosdaq_prices().get(nm)
        amt = round(ch * price) if price else None   # 변화금액(원) = 변화 × 종가
        rows.append({"종목명": nm, "당일수량": tq, "전일수량": pq, "변화": ch, "구분": typ, "변화금액": amt})
    rows.sort(key=lambda r: -abs(r["변화"]))
    return rows

def _fetch(e, ymd):
    """모드 무관 통일 호출 → (map, 기준일YYYYMMDD)."""
    if e["mode"] == "dateapi":
        return e["fetch"](e["id"], ymd)
    return e["fetch"](e["id"])  # snapshot: ymd 무시, 최신 반환


def run(today=None):
    """증분 구조: 이미 당일분 캡처된 ETF는 재요청 스킵, 미게시/지연만 매번 시도해 채움.
    매 실행 시 당일 파일(xlsx/csv)을 '현재까지 캡처된 전체 상태'로 덮어씀."""
    today = today or datetime.now().date()
    if not is_trading_day(today):
        print(f"[{today}] 휴장일 — 스킵"); return
    t_ymd = today.strftime("%Y%m%d")
    prev = prev_trading_day(today); p_ymd = prev.strftime("%Y%m%d")
    snap = load_snap()
    print(f"=== PDF 변화(증분): 당일 {today} vs 전영업일 {prev} ===")
    for e in ETFS:
        key = f"{e['am']}:{e['id']}"; snap.setdefault(key, {})
        # 1) 전일 베이스라인 (없을 때만 1회)
        if p_ymd not in snap[key]:
            try:
                pm, pa = _fetch(e, p_ymd)
                if pm and pa: snap[key][pa] = pm
            except Exception as ex:
                print(f"  [{e['name'][:16]}] 전일 조회 실패: {type(ex).__name__}")
        # 2) 당일 — 이미 있으면 스킵(증분), 없으면 시도
        if t_ymd in snap[key]:
            status = "이미 캡처(스킵)"
        else:
            try:
                tm, ta = _fetch(e, t_ymd)
                if tm and ta == t_ymd:
                    snap[key][t_ymd] = tm; status = f"신규 캡처 {len(tm)}종목"
                else:
                    status = f"대기 (현재 {ta or '빈값'})"
            except Exception as ex:
                status = f"대기 (오류 {type(ex).__name__})"
        ready = (t_ymd in snap[key]) and (p_ymd in snap[key])
        print(f"  {'✅' if ready else '⏳'} {e['name'][:18]:18} {status}")
    save_snap(snap)

    # 출력 빌드 (스냅샷에서, 8개 전부 상태 포함)
    groups = []; csv_rows = []; done = 0; pend = []
    for e in ETFS:
        kk = snap.get(f"{e['am']}:{e['id']}", {})
        if t_ymd in kk and p_ymd in kk:
            done += 1
            rows = [r for r in diff(kk[t_ymd], kk[p_ymd])
                    if r["구분"] != "유지"
                    and (r["구분"] in ("신규편입", "편출") or abs(r["변화"]) >= MIN_CHANGE)]
            nc = num_cu(e)
            stockval = sum(q * p for q, p in (_qp(v) for v in kk[t_ymd].values()) if p) * nc  # 보유주식총액(PDF, fallback)
            aum = etf_navtotal(e) or (round(stockval) if stockval else None)  # 비중 분모 = KRX 순자산총액(우선)
            for r in rows:
                r["전체금액"] = round(r["변화금액"] * nc) if (r["변화금액"] is not None and nc) else None
                r["비중"] = round(r["전체금액"] / aum * 100, 3) if (r.get("전체금액") is not None and aum) else None
            buy = sorted([r for r in rows if r["구분"] in ("신규편입", "수량확대")], key=lambda r: -r["변화"])
            sell = sorted([r for r in rows if r["구분"] in ("수량축소", "편출")], key=lambda r: r["변화"])
            groups.append({"etf": e["name"], "am": e["am"], "state": "captured", "buy": buy, "sell": sell, "aum": aum})
            for r in rows:
                csv_rows.append([e["name"], e["am"], t_ymd, p_ymd, r["구분"], r["종목명"],
                                 int(r["당일수량"]), int(r["전일수량"]), int(r["변화"]),
                                 (r["변화금액"] if r["변화금액"] is not None else ""),
                                 (r["전체금액"] if r["전체금액"] is not None else ""),
                                 (r["비중"] if r.get("비중") is not None else "")])
        else:
            groups.append({"etf": e["name"], "am": e["am"], "state": "pending", "buy": [], "sell": []})
            pend.append(e["name"].split()[0])
    status_line = f"캡처 {done}/{len(ETFS)}" + (f"  ·  대기: {', '.join(pend)}" if pend else "  ·  전부 반영 완료")
    csv_path = os.path.join(BASE, f"pdf_change_{t_ymd}.csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["ETF","운용사","당일기준일","전일기준일","구분","종목명","당일수량","전일수량","수량변화(1CU)","1CU변화금액(원)","전체매매금액(원)","전체대비비중(%)"])
        for r in csv_rows: w.writerow(r)
    xlsx_path = write_excel(groups, today, prev, status_line)
    print(f"\n[{status_line}]  변화 {len(csv_rows)}행 → {os.path.basename(xlsx_path)}")

    # 그 주 마지막 거래일(보통 금요일, 휴장이면 목요일)이면 주간 리포트도 생성
    if is_last_trading_day_of_week(today):
        try:
            weekly_report(today)
        except Exception as ex:
            print(f"[주간] 생성 실패: {type(ex).__name__}: {ex}")

    return {"groups": groups, "today": today, "prev": prev, "status_line": status_line,
            "csv_path": csv_path, "xlsx_path": xlsx_path, "done": done, "pending": pend}


def week_trading_days(any_day):
    """any_day가 속한 주의 거래일 리스트(과거→현재). 월요일부터 any_day까지 중 거래일만.
    (금요일 실행=그 주 전체, 주중 실행=그 주 진행분까지 = '이번주 현재까지')."""
    monday = any_day - timedelta(days=any_day.weekday())
    return [monday + timedelta(days=i) for i in range(any_day.weekday() + 1)
            if is_trading_day(monday + timedelta(days=i))]

def is_last_trading_day_of_week(d):
    """d가 그 주(월~금)의 마지막 거래일인가 (이후 금요일까지 거래일 없음)."""
    if not is_trading_day(d):
        return False
    fri = d - timedelta(days=d.weekday()) + timedelta(days=4)
    nxt = d + timedelta(days=1)
    while nxt <= fri:
        if is_trading_day(nxt):
            return False
        nxt += timedelta(days=1)
    return True

def nth_trading_day_before(d, n):
    x = d
    for _ in range(n):
        x = prev_trading_day(x)
    return x

def _build_period_groups(first, last):
    """first↔last PDF 비교 그룹 빌드. return (groups, done, pend_names)."""
    f_ymd, l_ymd = first.strftime("%Y%m%d"), last.strftime("%Y%m%d")
    groups = []; done = 0; pend = []
    for e in ETFS:
        try:
            fm, _ = _fetch(e, f_ymd); lm, _ = _fetch(e, l_ymd)
        except Exception:
            fm = lm = {}
        if fm and lm:
            done += 1
            rows = [r for r in diff(lm, fm)
                    if r["구분"] != "유지"
                    and (r["구분"] in ("신규편입", "편출") or abs(r["변화"]) >= MIN_CHANGE)]
            nc = num_cu(e)
            stockval = sum(q * p for q, p in (_qp(v) for v in lm.values()) if p) * nc  # 보유주식총액(마지막일, fallback)
            aum = etf_navtotal(e) or (round(stockval) if stockval else None)  # 비중 분모 = KRX 순자산총액(우선)
            for r in rows:
                r["전체금액"] = round(r["변화금액"] * nc) if (r["변화금액"] is not None and nc) else None
                r["비중"] = round(r["전체금액"] / aum * 100, 3) if (r.get("전체금액") is not None and aum) else None
            buy = sorted([r for r in rows if r["구분"] in ("신규편입", "수량확대")], key=lambda r: -r["변화"])
            sell = sorted([r for r in rows if r["구분"] in ("수량축소", "편출")], key=lambda r: r["변화"])
            groups.append({"etf": e["name"], "am": e["am"], "state": "captured", "buy": buy, "sell": sell, "aum": aum})
        else:
            groups.append({"etf": e["name"], "am": e["am"], "state": "pending", "buy": [], "sell": []})
            pend.append(e["name"].split()[0])
    return groups, done, pend

def weekly_report(asof=None):
    """그 주 첫 거래일 PDF ↔ 마지막 거래일 PDF 비교 = 주간 순변화."""
    asof = asof or datetime.now().date()
    tdays = week_trading_days(asof)
    if len(tdays) < 2:
        print(f"[주간] {asof} 주의 거래일 {len(tdays)}일 — 비교 불가"); return None
    first, last = tdays[0], tdays[-1]
    print(f"=== 주간 변화: {first}(주초) ↔ {last}(주말)  거래일 {len(tdays)}일 ===")
    groups, done, pend = _build_period_groups(first, last)
    span = f"{first.strftime('%Y-%m-%d')} ~ {last.strftime('%Y-%m-%d')}  ({len(tdays)}거래일)"
    status = (f"주간 변화  ·  {span}  ·  캡처 {done}/{len(ETFS)}" + (f"  ·  대기: {', '.join(pend)}" if pend else ""))
    path = write_excel(groups, last, first, status, title=f"코스닥 액티브 ETF 주간 PDF 변화   ·   {span}",
                       lbl_cur="주말", lbl_prev="주초", fname=f"pdf_weekly_{last.strftime('%Y%m%d')}.xlsx")
    print(f"[주간 저장] {os.path.basename(path)}")
    return path

def rolling_report(asof=None, ndays=5):
    """오늘 기준 최근 N거래일 순변화 = PDF(오늘) ↔ PDF(N거래일 전)."""
    asof = asof or datetime.now().date()
    while not is_trading_day(asof):
        asof = prev_trading_day(asof)
    last = asof
    first = nth_trading_day_before(last, ndays)
    print(f"=== 최근 {ndays}거래일 변화: {first} ↔ {last} ===")
    groups, done, pend = _build_period_groups(first, last)
    span = f"{first.strftime('%Y-%m-%d')} ~ {last.strftime('%Y-%m-%d')}  (최근 {ndays}거래일)"
    status = (f"최근 {ndays}거래일 변화  ·  {span}  ·  캡처 {done}/{len(ETFS)}" + (f"  ·  대기: {', '.join(pend)}" if pend else ""))
    path = write_excel(groups, last, first, status, title=f"코스닥 액티브 ETF 최근 {ndays}거래일 PDF 변화   ·   {span}",
                       lbl_cur=f"오늘", lbl_prev=f"{ndays}일전", fname=f"pdf_rolling{ndays}_{last.strftime('%Y%m%d')}.xlsx")
    print(f"[롤링 저장] {os.path.basename(path)}")
    return path


def write_excel(groups, today, prev, status_line="", title=None, lbl_cur="당일", lbl_prev="전영업일", fname=None):
    """8개 ETF 전부 표시(변화있음=좌매수/우매도, 변화없음, 대기). 날짜 YYYY-MM-DD.
    groups: [{etf, am, state('captured'/'pending'), buy:[diff행dict], sell:[...]}]
    title/lbl_cur/lbl_prev: 주간 등 다른 기간 리포트용 라벨 오버라이드."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "PDF변화"
    RED = Font(color="C00000", bold=True); BLUE = Font(color="0070C0", bold=True)
    BANNERF = Font(bold=True, size=11, color="1F3864"); BANNERBG = PatternFill("solid", fgColor="DCE6F2")
    PENDF = Font(bold=True, size=11, color="7F7F7F"); PENDBG = PatternFill("solid", fgColor="EDEDED")
    BUYHEAD = Font(bold=True, color="FFFFFF"); BUYHBG = PatternFill("solid", fgColor="C00000")
    SELLHEAD = Font(bold=True, color="FFFFFF"); SELLHBG = PatternFill("solid", fgColor="0070C0")
    BUYBG = PatternFill("solid", fgColor="FCE4E4"); SELLBG = PatternFill("solid", fgColor="E4ECF6")
    thin = Side(style="thin", color="D9D9D9"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    NUMFMT_UP = "+#,##0;-#,##0;0"   # 수량(주) 부호표시
    AMTFMT = "+#,##0;-#,##0;0"      # 변화금액(원) 부호표시
    PCTFMT = '+0.00"%";-0.00"%";0'  # 전체대비 비중 부호표시
    td = today.strftime("%Y-%m-%d"); pd_ = prev.strftime("%Y-%m-%d")
    NC = 11  # A~K (좌5 + gap + 우5)
    def mergerow(r): ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)

    mergerow(1)
    ws["A1"] = title or f"코스닥 액티브 ETF PDF 변화   ·   {lbl_cur} {td}  vs  {lbl_prev} {pd_}"
    ws["A1"].font = Font(bold=True, size=14); ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    if status_line:
        mergerow(2); ws["A2"] = status_line
        ws["A2"].font = Font(size=10, color="595959", italic=True)
    row = 3
    for g in groups:
        etf, am, state = g["etf"], g["am"], g["state"]
        buy = sorted(g["buy"], key=lambda r: -r["변화"]); sell = sorted(g["sell"], key=lambda r: r["변화"])
        mergerow(row)
        if state == "pending":
            bc = ws.cell(row, 1, f"■ {etf}  ({am})      ⏳ 대기 — 당일 PDF 미게시/지연 (게시되면 자동 반영)")
            bc.font = PENDF; bc.alignment = Alignment(vertical="center")
            for c in range(1, NC + 1): ws.cell(row, c).fill = PENDBG
            ws.row_dimensions[row].height = 19; row += 2; continue
        aum = g.get("aum")
        aum_txt = f"      순자산 ≈ {aum/1e8:,.0f}억" if aum else ""
        bc = ws.cell(row, 1, f"■ {etf}  ({am}){aum_txt}      {lbl_cur} {td}  vs  {lbl_prev} {pd_}      매수 {len(buy)}종목  /  매도 {len(sell)}종목")
        bc.font = BANNERF; bc.alignment = Alignment(vertical="center")
        for c in range(1, NC + 1): ws.cell(row, c).fill = BANNERBG
        ws.row_dimensions[row].height = 19; row += 1
        if not buy and not sell:
            mergerow(row)
            nc = ws.cell(row, 1, "변화 없음 (구성종목 동일)"); nc.font = Font(color="808080", italic=True)
            row += 2; continue
        # 좌(매수 1~5) / 우(매도 7~11) 헤더
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=11)
        hl = ws.cell(row, 1, "🔴 매수 (확대·신규편입)"); hl.font = BUYHEAD; hl.alignment = Alignment(horizontal="center")
        hr = ws.cell(row, 7, "🔵 매도 (축소·편출)"); hr.font = SELLHEAD; hr.alignment = Alignment(horizontal="center")
        for c in (1, 2, 3, 4, 5): ws.cell(row, c).fill = BUYHBG
        for c in (7, 8, 9, 10, 11): ws.cell(row, c).fill = SELLHBG
        row += 1
        for c, t in zip((1, 2, 3, 4, 5, 7, 8, 9, 10, 11),
                        ("종목명", "변화(1CU주)", "전체매매금액(원)", "전체대비비중", "전일→당일",
                         "종목명", "변화(1CU주)", "전체매매금액(원)", "전체대비비중", "전일→당일")):
            cell = ws.cell(row, c, t); cell.font = Font(bold=True, size=9, color="595959")
            cell.alignment = Alignment(horizontal="center"); cell.border = border
        row += 1
        def put(r0, base, fnt, bg):
            nm = r0["종목명"]; pq = int(r0["전일수량"]); tq = int(r0["당일수량"])
            amt = r0.get("전체금액"); pct = r0.get("비중")
            tag = "  (신규)" if r0["구분"] == "신규편입" else ("  (편출)" if r0["구분"] == "편출" else "")
            ws.cell(row, base, nm + tag).font = fnt
            cc = ws.cell(row, base + 1, int(r0["변화"])); cc.number_format = NUMFMT_UP; cc.font = fnt; cc.fill = bg; cc.alignment = Alignment(horizontal="right")
            ca = ws.cell(row, base + 2, amt if amt is not None else "-"); ca.font = fnt; ca.fill = bg; ca.alignment = Alignment(horizontal="right")
            if amt is not None: ca.number_format = AMTFMT
            cp = ws.cell(row, base + 3, pct if pct is not None else "-"); cp.font = fnt; cp.fill = bg; cp.alignment = Alignment(horizontal="right")
            if pct is not None: cp.number_format = PCTFMT
            ws.cell(row, base + 4, f"{pq:,}→{tq:,}").alignment = Alignment(horizontal="right")
        for i in range(max(len(buy), len(sell))):
            if i < len(buy): put(buy[i], 1, RED, BUYBG)
            if i < len(sell): put(sell[i], 7, BLUE, SELLBG)
            for c in (1, 2, 3, 4, 5, 7, 8, 9, 10, 11): ws.cell(row, c).border = border
            row += 1
        row += 1

    for col, w in zip("ABCDEFGHIJK", [18, 9, 14, 11, 13, 3, 18, 9, 14, 11, 13]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    base_name = fname or f"pdf_change_{today.strftime('%Y%m%d')}.xlsx"
    path = os.path.join(BASE, base_name)
    # 과거에 남았을 수 있는 _HHMM 대체파일 청소(하루 1개 유지)
    import glob as _glob
    stem = base_name.rsplit(".", 1)[0]
    for old in _glob.glob(os.path.join(BASE, f"{stem}_*.xlsx")):
        try: os.remove(old)
        except OSError: pass
    try:
        wb.save(path)
        return path
    except PermissionError:
        # 파일이 Excel에 열려 잠김 → 이번 회차 저장 스킵(대체파일 만들지 않음).
        # 다음 회차(10분 뒤)에 파일이 닫혀 있으면 동일 이름으로 덮어써짐 → 하루 1개 유지.
        print(f"  [스킵] {base_name} 열려있어 저장 보류 — 파일 닫으면 다음 회차에 갱신됨")
        return path

# ── 텔레그램용 이미지 렌더 + 전송 ─────────────────────────────
def _setup_mpl_font():
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.font_manager as fm
    avail = {f.name for f in fm.fontManager.ttflist}
    for f in ("Malgun Gothic", "NanumGothic", "NanumBarunGothic", "AppleGothic", "Noto Sans CJK KR"):
        if f in avail:
            matplotlib.rcParams["font.family"] = f; break
    matplotlib.rcParams["axes.unicode_minus"] = False

def _eok(x):
    """원 → '±NN.N억' 문자열(없으면 '-')."""
    return "-" if x is None else f"{x/1e8:+,.1f}억"

def render_report_image(groups, today, prev, status_line="", title=None,
                        lbl_cur="당일", lbl_prev="전영업일", out_png=None):
    """엑셀 표를 한 장의 PNG로 렌더(ETF별 매수/매도). 반환: 경로, 변화 없으면 None."""
    _setup_mpl_font()
    import matplotlib.pyplot as plt
    from matplotlib.gridspec import GridSpec
    caps = [g for g in groups if g["state"] == "captured" and (g["buy"] or g["sell"])]
    if not caps:
        return None
    td = today.strftime("%Y-%m-%d"); pdd = prev.strftime("%Y-%m-%d")
    nrows = [max(len(g["buy"]), len(g["sell"])) for g in caps]
    ratios = [n + 1.6 for n in nrows]
    fig_h = 1.0 + sum(r * 0.32 for r in ratios)
    fig = plt.figure(figsize=(11.5, fig_h), dpi=170)
    gs = GridSpec(len(caps), 1, height_ratios=ratios, hspace=0.6)
    RED = "#C00000"; BLUE = "#0070C0"; RBG = "#FCE4E4"; BBG = "#E4ECF6"
    for gi, g in enumerate(caps):
        ax = fig.add_subplot(gs[gi]); ax.axis("off")
        buy = sorted(g["buy"], key=lambda r: -r["변화"]); sell = sorted(g["sell"], key=lambda r: r["변화"])
        aum = g.get("aum")
        ttl = f"■ {g['etf']}  ({g['am']})" + (f"   순자산 약 {aum/1e8:,.0f}억" if aum else "") + \
              f"      매수 {len(buy)}  /  매도 {len(sell)}"
        ax.set_title(ttl, loc="left", fontsize=11, fontweight="bold", color="#1F3864", pad=6)
        n = max(len(buy), len(sell))
        cols = ["매수 종목", "변화(주)", "금액", "비중", "", "매도 종목", "변화(주)", "금액", "비중"]
        cell = []; ccol = []
        for i in range(n):
            row = [""] * 9; col = ["white"] * 9
            if i < len(buy):
                b = buy[i]; tag = " (신규)" if b["구분"] == "신규편입" else ""
                row[0] = b["종목명"] + tag; row[1] = f"{int(b['변화']):+,}"; row[2] = _eok(b.get("전체금액"))
                row[3] = f"{b['비중']:+.2f}%" if b.get("비중") is not None else "-"
                for c in (0, 1, 2, 3): col[c] = RBG
            if i < len(sell):
                s = sell[i]; tag = " (편출)" if s["구분"] == "편출" else ""
                row[5] = s["종목명"] + tag; row[6] = f"{int(s['변화']):+,}"; row[7] = _eok(s.get("전체금액"))
                row[8] = f"{s['비중']:+.2f}%" if s.get("비중") is not None else "-"
                for c in (5, 6, 7, 8): col[c] = BBG
            cell.append(row); ccol.append(col)
        tbl = ax.table(cellText=cell, colLabels=cols, cellColours=ccol,
                       colColours=[RED, RED, RED, RED, "white", BLUE, BLUE, BLUE, BLUE],
                       cellLoc="center", loc="upper center",
                       colWidths=[0.23, 0.09, 0.10, 0.09, 0.02, 0.23, 0.09, 0.10, 0.09])
        tbl.auto_set_font_size(False); tbl.set_fontsize(8.5); tbl.scale(1, 1.28)
        for c in range(9):
            if c != 4:
                t = tbl[0, c].get_text(); t.set_color("white"); t.set_fontweight("bold")
    sub = f"\n{status_line}" if status_line else ""
    sup = title or f"코스닥 액티브 ETF PDF 변화   ·   {lbl_cur} {td}  vs  {lbl_prev} {pdd}"
    fig.suptitle(sup + sub, fontsize=13, fontweight="bold", y=0.999)
    out_png = out_png or os.path.join(BASE, f"pdf_change_{today.strftime('%Y%m%d')}.png")
    fig.savefig(out_png, bbox_inches="tight", facecolor="white"); plt.close(fig)
    return out_png

def send_telegram(png_path, caption=""):
    """TELEGRAM_TOKEN/TELEGRAM_CHAT_ID 있으면 이미지 전송(sendPhoto→실패 시 sendDocument)."""
    tok = os.environ.get("TELEGRAM_TOKEN"); chat = os.environ.get("TELEGRAM_CHAT_ID")
    if not (tok and chat):
        print("  [텔레그램] 토큰/챗ID 미설정 — 전송 스킵"); return False
    if not (png_path and os.path.exists(png_path)):
        print("  [텔레그램] 보낼 이미지 없음"); return False
    for method, field in (("sendPhoto", "photo"), ("sendDocument", "document")):
        try:
            with open(png_path, "rb") as f:
                r = requests.post(f"https://api.telegram.org/bot{tok}/{method}",
                                  data={"chat_id": chat, "caption": caption[:1024]},
                                  files={field: f}, timeout=60)
            if r.ok and r.json().get("ok"):
                print(f"  [텔레그램] {method} 전송 완료"); return True
            print(f"  [텔레그램] {method} 실패: {r.text[:150]}")
        except Exception as e:
            print(f"  [텔레그램] {method} 오류: {type(e).__name__}: {e}")
    return False


if __name__ == "__main__":
    args = sys.argv[1:]
    if args and args[0] == "--weekly":
        d = datetime.strptime(args[1], "%Y-%m-%d").date() if len(args) > 1 else None
        weekly_report(d)
    elif args and args[0] == "--days":   # 최근 N거래일 롤링: --days 5 [YYYY-MM-DD]
        n = int(args[1]) if len(args) > 1 else 5
        d = datetime.strptime(args[2], "%Y-%m-%d").date() if len(args) > 2 else None
        rolling_report(d, n)
    else:
        d = datetime.strptime(args[0], "%Y-%m-%d").date() if args else None
        run(d)
